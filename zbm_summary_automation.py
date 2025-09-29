import pandas as pd
import numpy as np
from datetime import datetime

def create_zbm_summary():
    """
    Automate the creation of ZBM summary sheet from master_tracker.csv
    """
    
    print("🔄 Starting ZBM Summary Automation...")
    
    # Read master tracker data
    print("📖 Reading master_tracker.csv...")
    try:
        # Try different encodings
        encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        df = None
        
        for encoding in encodings:
            try:
                df = pd.read_csv('master_tracker.csv', encoding=encoding)
                print(f"✅ Successfully loaded {len(df)} records from master_tracker.csv using {encoding} encoding")
                break
            except UnicodeDecodeError:
                continue
        
        if df is None:
            print("❌ Could not read master_tracker.csv with any of the tried encodings")
            return
            
    except Exception as e:
        print(f"❌ Error reading master_tracker.csv: {e}")
        return
    
    # Clean and prepare data
    print("🧹 Cleaning and preparing data...")
    
    # Ensure required columns exist
    required_columns = ['ABM Terr Code', 'TBM HQ', 'ABM Name',
                        'Doctor: Customer Code', 'Assigned Request Ids', 'Request Status']
    missing = [c for c in required_columns if c not in df.columns]
    if missing:
        print(f"❌ Missing required columns in master_tracker.csv: {missing}")
        return

    # Remove rows where key fields are null or empty
    df = df.dropna(subset=['ABM Terr Code', 'TBM HQ', 'ABM Name'])
    df = df[df['ABM Terr Code'].astype(str).str.strip() != '']
    df = df[df['TBM HQ'].astype(str).str.strip() != '']
    df = df[df['ABM Name'].astype(str).str.strip() != '']

    # Restrict to specified TBM HQ cities
    allowed_hq = {"MUMBAI", "AHMEDABAD", "PUNE", "NAGPUR"}
    df['TBM HQ'] = df['TBM HQ'].astype(str)
    df = df[df['TBM HQ'].str.upper().isin(allowed_hq)]

    # Build Area Name: "TBM HQ - ABM Terr Code"
    df['Area Name'] = (
        df['TBM HQ'].astype(str).str.strip() + ' - ' + df['ABM Terr Code'].astype(str).str.strip()
    )

    # Compute Final Answer per unique request id using rules from logic.xlsx
    print("🧠 Computing final status per unique Request Id using rules (test.py logic)...")
    try:
        xls_rules = pd.ExcelFile('logic.xlsx')
        sheet2 = pd.read_excel(xls_rules, 'Sheet2')

        def normalize(text):
            return str(text).strip().casefold()

        rules = {}
        for _, row in sheet2.iterrows():
            statuses = [normalize(s) for s in row.drop('Final Answer').dropna().tolist()]
            statuses = tuple(sorted(set(statuses)))
            rules[statuses] = row['Final Answer']

        # Group statuses by request id from master data
        grouped = df.groupby('Assigned Request Ids')['Request Status'].apply(list).reset_index()

        def get_final_answer(status_list):
            key = tuple(sorted(set(normalize(s) for s in status_list)))
            return rules.get(key, '❌ No matching rule')

        grouped['Request Status'] = grouped['Request Status'].apply(lambda lst: sorted(set(lst), key=str))
        grouped['Final Answer'] = grouped['Request Status'].apply(get_final_answer)

        def has_action_pending(status_list):
            target = 'action pending / in process'
            return any(normalize(s) == target for s in status_list)
        grouped['Has D Pending'] = grouped['Request Status'].apply(has_action_pending)

        # Try to save with timestamp to avoid permission issues
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_file = f'final_output_{timestamp}.xlsx'
        
        try:
            grouped.to_excel(output_file, index=False)
            print(f"✅ Final Answer per request saved to {output_file}")
        except PermissionError:
            print("⚠️ Permission denied for Excel file. Trying alternative approach...")
            # Save as CSV instead
            csv_output = f'final_output_{timestamp}.csv'
            grouped.to_csv(csv_output, index=False)
            print(f"✅ Final Answer per request saved as CSV: {csv_output}")
            print("💡 Note: Excel file may be open. Close it and run again for Excel format.")

        # Merge Final Answer back to main dataframe
        df = df.merge(grouped[['Assigned Request Ids', 'Final Answer']], on='Assigned Request Ids', how='left')
    except Exception as e:
        print(f"❌ Error computing final status from logic.xlsx: {e}")
        return
    
    print(f"📊 After cleaning: {len(df)} records remaining")
    
    # Group by ABM Terr Code and ABM Name
    print("📈 Aggregating data by ABM Terr Code...")
    
    tbm_code_col = 'TBM Terr Code' if 'TBM Terr Code' in df.columns else 'ABM Terr Code'

    aggregated = df.groupby(['Area Name', 'ABM Name']).agg({
        tbm_code_col: 'nunique',             # Unique TBMs
        'Doctor: Customer Code': 'nunique',  # Unique HCPs
        'Assigned Request Ids': 'nunique',   # Unique Requests
    }).reset_index()
    
    aggregated = aggregated.rename(columns={
        tbm_code_col: 'Unique TBMs',
        'Doctor: Customer Code': 'Unique HCPs',
        'Assigned Request Ids': 'Unique Requests',
    })
    
    # Define status categories
    status_categories = {
        'out_of_stock_on_hold': ['Out of stock', 'On hold', 'Not permitted'],
        'request_raised': ['Request Raised'],
        'delivered_return_action_pending': ['Delivered', 'Return', 'Action pending / In Process', 'Dispatched & In Transit', 'Dispatch Pending'],
        'action_pending': ['Action pending / In Process'],
        'dispatch_pending': ['Dispatch Pending'],
        'delivered': ['Delivered'],
        'dispatched_in_transit': ['Dispatched & In Transit'],
        'rto': ['RTO']
    }
    
    print("🔢 Calculating status-specific metrics...")

    # ✅ FIX: count unique request IDs, not rows
    for category_name, status_list in status_categories.items():
        print(f"   Processing {category_name}...")

        status_counts = df.groupby(['Area Name', 'ABM Name']).apply(
            lambda x: x.loc[x['Final Answer'].isin(status_list), 'Assigned Request Ids'].nunique()
        ).reset_index(name=f'count_{category_name}')

        aggregated = aggregated.merge(status_counts, on=['Area Name', 'ABM Name'], how='left')
    
    # Derived metrics following template formulas exactly
    print("📊 Calculating derived metrics following template formulas...")

    # HO metrics (A + B)
    aggregated['Request Cancelled Out of Stock'] = aggregated['count_out_of_stock_on_hold']  # A
    aggregated['Action Pending at HO'] = aggregated['count_action_pending']  # B
    
    # HUB metrics (D + E + F)
    d_counts = df.merge(grouped[['Assigned Request Ids', 'Has D Pending']], on='Assigned Request Ids', how='left') \
                .groupby(['Area Name', 'ABM Name']) \
                .apply(lambda x: x.loc[x['Has D Pending'] == True, 'Assigned Request Ids'].nunique()) \
                .reset_index(name='Pending for Invoicing')
    aggregated = aggregated.merge(d_counts, on=['Area Name', 'ABM Name'], how='left')
    aggregated['Pending for Invoicing'] = aggregated['Pending for Invoicing'].fillna(0).astype(int)  # D
    
    aggregated['Pending for Dispatch'] = aggregated['count_dispatch_pending']  # E
    
    # Delivery Status (G + H + I)
    aggregated['Delivered'] = aggregated['count_delivered']  # G
    aggregated['Dispatched In Transit'] = aggregated['count_dispatched_in_transit']  # H
    aggregated['RTO'] = aggregated['count_rto']  # I
    
    # Calculate derived metrics using template formulas
    aggregated['Requests Dispatched'] = aggregated['Delivered'] + aggregated['Dispatched In Transit'] + aggregated['RTO']  # F = G + H + I
    aggregated['Sent to HUB'] = aggregated['Pending for Invoicing'] + aggregated['Pending for Dispatch'] + aggregated['Requests Dispatched']  # C = D + E + F
    aggregated['Requests Raised'] = aggregated['Request Cancelled Out of Stock'] + aggregated['Action Pending at HO'] + aggregated['Sent to HUB']  # A + B + C
    
    # RTO Reasons (placeholders)
    aggregated['Incomplete Address'] = 0
    aggregated['Doctor Non Contactable'] = 0
    aggregated['Doctor Refused to Accept'] = 0
    aggregated['Hold Delivery'] = 0
    
    # Final summary
    print("📋 Creating final summary...")

    summary_columns = [
        'Area Name', 'ABM Name', 'Unique TBMs', 'Unique HCPs', 'Unique Requests', 'Requests Raised',
        'Request Cancelled Out of Stock', 'Action Pending at HO', 'Sent to HUB',
        'Pending for Invoicing', 'Pending for Dispatch', 'Requests Dispatched',
        'Delivered', 'Dispatched In Transit', 'RTO',
        'Incomplete Address', 'Doctor Non Contactable', 'Doctor Refused to Accept', 'Hold Delivery'
    ]
    
    final_summary = aggregated[summary_columns].copy()
    final_summary = final_summary.sort_values('Area Name').reset_index(drop=True)
    
    csv_output = f"zbm_summary_output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    final_summary.to_csv(csv_output, index=False)
    print(f"💾 Saved summary data to {csv_output} for verification")
    
    # Write values into template ZBM sheet with exact format matching
    print("💾 Writing values into template ZBM sheet (exact format matching)...")
    
    try:
        from openpyxl import load_workbook
        from copy import copy as copy_style

        wb = load_workbook('zbm_summary.xlsx')
        ws = wb['ZBM']

        # Preserve original sheet formatting
        print("📋 Preserving original template formatting...")
        
        # Protect template structure - don't modify rows 1-3 (headers)
        data_start_row = 3  # Data starts from row 4 (index 3)
        max_clear_rows = max(len(final_summary) + 10, 100)
        
        # Clear only data area (columns B to S, rows 4 onwards) - preserve headers
        for r in range(data_start_row + 1, data_start_row + max_clear_rows):
            for c in range(2, 20):  # Columns B to S
                ws.cell(row=r, column=c).value = None

        # Define exact column mapping based on template analysis
        column_mapping = {
            'Area Name': 2,           # Column B - ABM Terr Code
            'ABM Name': 3,           # Column C - Input Sample Request: Created By
            'Unique TBMs': 4,        # Column D - Doctor: Customer Code (should be NaN in template)
            'Unique HCPs': 5,        # Column E - unique count of (Doctor: Customer Code)
            'Requests Raised': 6,    # Column F - unique count of (Assigned Request Ids)
            'Request Cancelled Out of Stock': 7,  # Column G - HO section
            'Action Pending at HO': 8,            # Column H - HO section
            'Sent to HUB': 9,                    # Column I - HUB section
            'Pending for Invoicing': 10,         # Column J - HUB section
            'Pending for Dispatch': 11,          # Column K - HUB section
            'Requests Dispatched': 12,           # Column L - Delivery Status
            'Delivered': 13,                     # Column M - Delivery Status
            'Dispatched In Transit': 14,         # Column N - Delivery Status
            'RTO': 15,                           # Column O - Delivery Status
            'Incomplete Address': 16,            # Column P - RTO Reasons
            'Doctor Non Contactable': 17,        # Column Q - RTO Reasons
            'Doctor Refused to Accept': 18,      # Column R - RTO Reasons
            'Hold Delivery': 19                  # Column S - RTO Reasons
        }

        def copy_row_style(src_row_idx, dst_row_idx):
            """Copy formatting from source row to destination row"""
            for c in range(2, 20):  # Columns B to S
                src = ws.cell(row=src_row_idx, column=c)
                dst = ws.cell(row=dst_row_idx, column=c)
                dst.number_format = src.number_format
                # Preserve original font formatting exactly
                dst.font = copy_style(src.font)
                dst.alignment = copy_style(src.alignment)
                dst.border = copy_style(src.border)
                dst.fill = copy_style(src.fill)

        # Write data rows
        for i in range(len(final_summary)):
            target_row = data_start_row + 1 + i  # Start from row 4
            if target_row > ws.max_row:
                ws.insert_rows(target_row)
            
            # Copy formatting from template row 4
            copy_row_style(4, target_row)
            
            # Write data according to exact column mapping
            for col_name, col_num in column_mapping.items():
                if col_name in final_summary.columns:
                    value = final_summary.at[i, col_name]
                    ws.cell(row=target_row, column=col_num).value = value
            
            # Set Column 4 (# Unique TBMs) to NaN as per template
            ws.cell(row=target_row, column=4).value = None

        # Add total row
        total_row = data_start_row + 1 + len(final_summary)
        if total_row > ws.max_row:
            ws.insert_rows(total_row)
        
        # Copy formatting for total row
        copy_row_style(4, total_row)
        
        # Write totals
        ws.cell(row=total_row, column=2).value = None  # Empty first column
        ws.cell(row=total_row, column=3).value = "Total"
        
        # Calculate and write totals for each column
        for col_name, col_num in column_mapping.items():
            if col_name in final_summary.columns and col_name not in ['Area Name', 'ABM Name']:
                total_value = final_summary[col_name].sum()
                ws.cell(row=total_row, column=col_num).value = total_value
        
        # Set Column 4 (# Unique TBMs) to NaN in total row as per template
        ws.cell(row=total_row, column=4).value = None

        new_excel_file = f"zbm_summary_updated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        try:
            wb.save(new_excel_file)
            print(f"✅ Successfully created updated Excel file with formatting preserved: {new_excel_file}")
        except PermissionError:
            print("⚠️ Permission denied for Excel file. Trying alternative approach...")
            # Save as CSV instead
            csv_output = f"zbm_summary_updated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            final_summary.to_csv(csv_output, index=False)
            print(f"✅ Summary data saved as CSV: {csv_output}")
            print("💡 Note: Excel file may be open. Close it and run again for Excel format.")
        
        print("\n📊 Summary Statistics:")
        print(f"   Total ABM Territories: {len(final_summary)}")
        print(f"   Total Unique HCPs: {final_summary['Unique HCPs'].sum()}")
        print(f"   Total Unique Requests: {final_summary['Unique Requests'].sum()}")
        print(f"   Total Delivered: {final_summary['Delivered'].sum()}")
        print(f"   Total RTO: {final_summary['RTO'].sum()}")
        
        print("\n📋 Sample of generated data:")
        print(final_summary.head(10).to_string(index=False))
        
    except Exception as e:
        print(f"❌ Error updating Excel file: {e}")
        return
    
    print("\n🎉 ZBM Summary automation completed successfully!")

def watch_and_build():
    import time
    import os
    path = 'master_tracker.csv'
    try:
        last_mtime = os.path.getmtime(path)
    except OSError:
        print("❌ master_tracker.csv not found for watch mode")
        return
    print("👀 Watching master_tracker.csv for changes. Press Ctrl+C to stop.")
    while True:
        try:
            time.sleep(2)
            mtime = os.path.getmtime(path)
            if mtime != last_mtime:
                print("🔁 Change detected. Rebuilding report...")
                last_mtime = mtime
                create_zbm_summary()
        except KeyboardInterrupt:
            print("👋 Watcher stopped.")
            break

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description='Generate ZBM summary or watch for changes')
    parser.add_argument('--watch', action='store_true', help='Watch master_tracker.csv and regenerate on change')
    args = parser.parse_args()
    if args.watch:
        watch_and_build()
    else:
        create_zbm_summary()
