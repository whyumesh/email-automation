import pandas as pd
import os
from jinja2 import Environment, FileSystemLoader
import win32com.client as win32
from datetime import datetime as dt
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.workbook.protection import WorkbookProtection
from win32com.client import Dispatch
import win32com.client as win32

from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.image as mpimg
from openpyxl import load_workbook

import pandas as pd
from openpyxl.utils import get_column_letter
from xlsxwriter.utility import xl_range
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

from datetime import datetime as dt
import datetime

# Load Jinja2 Template
env = Environment(loader=FileSystemLoader('.'))
template = env.get_template("email_template_ABM_V4.html")

#pathc = "C:\\Users\\PAGARAX1\\OneDrive - Abbott\\Documents\\Project 3D email automation\\"
#df = pd.read_excel(os.path.join(pathc, 'Sample Request - Master File.xlsx'))

#df = pd.read_excel('Sap Master Dummy File.xlsx')

df = pd.read_excel('TBM & ABM Automation Email 1209252.xlsx')

emp_codes = [730310,738595]
#df = df[df['ABM Emp Code'].isin(emp_codes)]

emp_codes_ZBM =[737217,52158]
#df = df[df['ZBM Emp Code'].isin(emp_codes_ZBM)]

#df['Date'] = df['Date'].dt.date

"""
df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
df['Date'] = df['Date'].dt.strftime('%d/%m/%Y')
z = dt.today()
current_date = z.date()

df['Delivery Date'] = pd.to_datetime(df['Delivery Date'], errors='coerce')
df['Delivery Date'] = df['Delivery Date'].dt.strftime('%d/%m/%Y')
df['Delivery Date'].fillna('-', inplace=True)
"""

df.rename(columns={'Request Date':'Date'}, inplace=True)
df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
df['Date'] = df['Date'].dt.strftime('%d-%b-%Y')
z = dt.today()
current_date = z.date()

df['Delivery Date'] = pd.to_datetime(df['Delivery Date'], errors='coerce')
df['Delivery Date'] = df['Delivery Date'].dt.strftime('%d-%b-%Y')
df['Delivery Date'].fillna('-', inplace=True)

df['Dispatch Date'] = pd.to_datetime(df['Dispatch Date'], errors='coerce')
df['Dispatch Date'] = df['Dispatch Date'].dt.strftime('%d-%b-%Y')
df['Dispatch Date'].fillna('-', inplace=True)


df['Rto Reason'].fillna('-', inplace=True)
df['Docket Number'].fillna('-', inplace=True)
df['Transporter Name'].fillna('-', inplace=True)

output_dir = os.path.dirname(os.path.abspath(__file__))

# Create the new folder path
new_folder_path = os.path.join(output_dir, f'ABM_files_{current_date}\\')

# Create the new folder
os.makedirs(new_folder_path, exist_ok=True)

#input_dir = "C:\\Users\\PAGARAX1\\Documents\\Consent Mail Automation\\test\\"

def excel_column_name(n):
    """Convert a zero-indexed column number to an Excel-style column name."""
    result = ''
    while n >= 0:
        result = chr(n % 26 + ord('A')) + result
        n = n // 26 - 1
    return result

# Function to save DataFrame as Excel file
def save_excel(df, filename):
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Write the DataFrame to the worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Define the table range
    table_range = f"A1:{chr(65 + df.shape[1] - 1)}{df.shape[0] + 1}"

    # Create a table
    table = Table(displayName="Table1", ref=table_range)

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(
        name="TableStyleLight13", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style

    # Add the table to the worksheet
    ws.add_table(table)
    
    
    # Autofit column width
    for i, column in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass

        adjusted_width = (max_length + 2)
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Set alignment for all cells
    for row in ws.iter_rows(min_row=1, max_row=df.shape[0] + 1, min_col=1, max_col=df.shape[1]):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')
    

    # Save the workbook
    wb.save(filename)

# Initialize Outlook
outlook = win32.Dispatch("Outlook.Application")

df['ABM Emp Code'] = df['ABM Emp Code'].astype(str)

df = df[(df['ABM EMAIL_ID'] != 0)&(df['ABM EMAIL_ID'] != '0')]
df = df[df['ABM EMAIL_ID'].notna()]

# Group by Created Alias
grouped = df.groupby("ABM Emp Code")
 
# Split DataFrame by area code and send emails
for alias, group in grouped:
    #email = group['email'].iloc[0]
    ABM_email = group['ABM EMAIL_ID'].iloc[0]
    ABM_name = group['ABM Name'].iloc[0]
    ABM_code = group['ABM Terr Code'].iloc[0]
    ZBM_email = group['ZBM EMAIL_ID'].iloc[0]
    filename = os.path.join(new_folder_path, f'Sample Request Raised - {ABM_code}.xlsx')
    #pdfname = os.path.join(input_dir, 'E-Consent Process-Project Reach.pdf')
    #password = 'abbott@123'  # Set your password here
    
    group.rename(columns={'Input Sample Request: Created By':'TBM Name',
                          'Date':'Requested_Date',
                          'Assigned Request Ids':'Request_ID',
                          'Doctor: Customer Code':'Doctor_Code',
                          'Doctor: SAP Customer Code(New)':'SAP_Customer_Code',
                          'Doctor: Account Name':'Doctor_Name',
                          'Item Code':'Item_Code',
                          'SKU':'SKU_Name',
                          'Requested Quantity':'Requested_Quantity',
                          'Request Status':'Request_Status',
                          'Delivery Date':'Delivery Date'}, inplace=True)
    
    group = group.loc[:,['AFFILIATE','TBM Division','TBM HQ','TBM Name','Requested_Date','Month','Request_ID', 'Doctor_Code', 'SAP_Customer_Code', 'Doctor_Name', 'Item_Code', 'SKU_Name', 'Requested_Quantity', 'Request_Status','Dispatch Date','Delivery Date','Rto Reason','Docket Number','Transporter Name']]
    
    #group = group.drop(columns=['email'])
    #group = group.drop(columns=['ABM_email'])
    #group = group.drop(columns=['ZBM_email'])
    
    # Save the DataFrame as an Excel file
    save_excel(group, filename)
    
 
    # Create and send email
    mail = outlook.CreateItem(0)
    mail.To = ABM_email
    mail.Attachments.Add(filename)
    if pd.notna(ZBM_email) and ZBM_email != 0:
        #print('yes')
        mail.cc = ZBM_email
    #mail.bCC = 
    mail.Subject = f"Sample Direct Dispatch to Doctors - Request Status as of {current_date}"
    mail.HTMLBody = template.render()
    mail.SentOnBehalfOfName = 'EPD_SFA@abbott.com'
    mail.Send()  # use mail.Send() to send automatically
       
print("All emails sent successfully!")    
    
    
    










