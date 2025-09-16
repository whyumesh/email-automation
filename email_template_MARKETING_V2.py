import pandas as pd
import os
from jinja2 import Environment, FileSystemLoader
import win32com.client as win32
from datetime import datetime as dt
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.workbook.protection import WorkbookProtection
from win32com.client import Dispatch
import win32com.client as win32

from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import NamedStyle
from openpyxl.utils.dataframe import dataframe_to_rows
import matplotlib.image as mpimg
from openpyxl import load_workbook

import pandas as pd
from openpyxl.utils import get_column_letter
from xlsxwriter.utility import xl_range
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook

from datetime import datetime as dt
import datetime

# Load Jinja2 Template
env = Environment(loader=FileSystemLoader('.'))
template = env.get_template("email_template_MARKETING - Tabular.html")

#pathc = "C:\\Users\\PAGARAX1\\OneDrive - Abbott\\Documents\\Project 3D email automation\\"
#df = pd.read_excel(os.path.join(pathc, 'Sample Request - Master File.xlsx'))

#df = pd.read_excel('Sap Master Dummy File.xlsx')

df = pd.read_csv("Marketing Head - ( Division Wise) 10092025.csv")

email_df = pd.read_excel("Affiliate wise Email I'D.xlsx")

emp_codes = [714810,737248]
#df = df[df['ABM Emp Code'].isin(emp_codes)]

emp_codes_ZBM =[737217,52158]
#df = df[df['ZBM Emp Code'].isin(emp_codes_ZBM)]

#df['Date'] = df['Date'].dt.date
df.rename(columns={'Request Date':'Date'}, inplace=True)
df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
df['Date'] = df['Date'].dt.strftime('%d-%b-%Y')
z = dt.today()
current_date = z.date()

df['Delivery Date'] = pd.to_datetime(df['Delivery Date'], errors='coerce')
df['Delivery Date'] = df['Delivery Date'].dt.strftime('%d-%b-%Y')
df['Delivery Date'].fillna('-', inplace=True)

df['Dispatch Date'] = pd.to_datetime(df['Dispatch Date'], errors='coerce')
df['Dispatch Date'] = df['Dispatch Date'].dt.strftime('%d-%b-%Y')
df['Dispatch Date'].fillna('-', inplace=True)


output_dir = os.path.dirname(os.path.abspath(__file__))

# Create the new folder path
new_folder_path = os.path.join(output_dir, f'MARKETING_files_{current_date}\\')

# Create the new folder
os.makedirs(new_folder_path, exist_ok=True)

#input_dir = "C:\\Users\\PAGARAX1\\Documents\\Consent Mail Automation\\test\\"

def excel_column_name(n):
    """Convert a zero-indexed column number to an Excel-style column name."""
    result = ''
    while n >= 0:
        result = chr(n % 26 + ord('A')) + result
        n = n // 26 - 1
    return result

# Function to save DataFrame as Excel file
def save_excel(df, filename):
    # Create a workbook and add a worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"

    # Write the DataFrame to the worksheet
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Define the table range
    table_range = f"A1:{chr(65 + df.shape[1] - 1)}{df.shape[0] + 1}"

    # Create a table
    table = Table(displayName="Table1", ref=table_range)

    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(
        name="TableStyleLight13", showFirstColumn=False,
        showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    table.tableStyleInfo = style

    # Add the table to the worksheet
    ws.add_table(table)
    
    
    # Autofit column width
    for i, column in enumerate(ws.columns, start=1):
        max_length = 0
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass

        adjusted_width = (max_length + 2)
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = adjusted_width

    # Set alignment for all cells
    for row in ws.iter_rows(min_row=1, max_row=df.shape[0] + 1, min_col=1, max_col=df.shape[1]):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')
    

    # Save the workbook
    wb.save(filename)
    
    


base_data = df.copy()



def complete_pivot_columns(df, col_used_f_pivot):
    
    ###################### Count of Requests dispatched ######################

    valid_requests = base_data[
        (base_data['Request Status'].notna()) &
        (~base_data['Request Status'].isin(['On hold','Request Raised','Action pending / In Process']))
    ]

    df['Count of Requests dispatched'] = df[f'{col_used_f_pivot}'].map(
        valid_requests.groupby(f'{col_used_f_pivot}')['Assigned Request Ids'].nunique()
    ).fillna(0).astype(int)


    ###################### Count of Requests Action pending / In Process ######################


    in_process_requests = base_data[(base_data['Request Status'].isin(['Action pending / In Process']))]


    df['Count of Requests - Action pending / In Process'] = df[f'{col_used_f_pivot}'].map(
        in_process_requests.groupby(f'{col_used_f_pivot}')['Assigned Request Ids'].nunique()
    ).fillna(0).astype(int)


    ###################### Count of Requests Dispatched & In Transit ######################

    in_transit_requests = base_data[(base_data['Request Status'].isin(['Dispatched & In Transit']))]


    df['Count of Requests - Dispatched & In Transit'] = df[f'{col_used_f_pivot}'].map(
        in_transit_requests.groupby(f'{col_used_f_pivot}')['Assigned Request Ids'].nunique()
    ).fillna(0).astype(int)

    ###################### Count of Out of stock ######################

    out_of_stock_requests = base_data[(base_data['Request Status'].isin(['Out of stock']))]


    df['Count of Requests - Out of stock'] = df[f'{col_used_f_pivot}'].map(
        out_of_stock_requests.groupby(f'{col_used_f_pivot}')['Assigned Request Ids'].nunique()
    ).fillna(0).astype(int)

    ###################### Count of Delivered  ######################

    delivered_requests = base_data[(base_data['Request Status'].isin(['Delivered']))]


    df['Count of Requests - Delivered'] = df[f'{col_used_f_pivot}'].map(
        delivered_requests.groupby(f'{col_used_f_pivot}')['Assigned Request Ids'].nunique()
    ).fillna(0).astype(int)

    ###################### Count of Return  ######################

    returned_requests = base_data[(base_data['Request Status'].isin(['Return']))]


    df['Count of Requests - Return'] = df[f'{col_used_f_pivot}'].map(
        returned_requests.groupby(f'{col_used_f_pivot}')['Assigned Request Ids'].nunique()
    ).fillna(0).astype(int)

    ###################### Count of Incomplete Address  ######################

    incomplete_addresses = base_data[(base_data['Rto Reason'].isin([' - Incomplete Address']))]


    df['Count of Requests - Incomplete Address'] = df[f'{col_used_f_pivot}'].map(
        incomplete_addresses.groupby(f'{col_used_f_pivot}')['Assigned Request Ids'].nunique()
    ).fillna(0).astype(int)


    ###################### Count of non contactable HCPs  ######################

    non_contactable_hcps = base_data[(base_data['Rto Reason'].isin([' - Dr. Non contactable']))]


    df['Count of Requests - Dr. Non contactable'] = df[f'{col_used_f_pivot}'].map(
        non_contactable_hcps.groupby(f'{col_used_f_pivot}')['Assigned Request Ids'].nunique()
    ).fillna(0).astype(int)


    ###################### Count of HCPs who refused to accept  ######################

    hcps_who_refused = base_data[(base_data['Rto Reason'].isin([' - Doctor refused to accept']))]


    df['Count of Requests - Doctor refused to accept'] = df[f'{col_used_f_pivot}'].map(
        hcps_who_refused.groupby(f'{col_used_f_pivot}')['Assigned Request Ids'].nunique()
    ).fillna(0).astype(int)

    
    
    return df


############### Division Level ###############

base_data_pivot = base_data.pivot_table(index=['AFFILIATE','TBM Division','DIV_NAME'], values=['Input Sample Request: Created Alias','Doctor: Customer Code','Assigned Request Ids'], aggfunc='nunique').reset_index()

base_data_pivot.rename(columns={'Input Sample Request: Created Alias':'Count of TBMs',
                                'Doctor: Customer Code':'Count of Unique HCPs Participating',
                                'Assigned Request Ids':'Count of Requests raised'}, inplace=True)


exst_cols = base_data_pivot.columns.tolist()[:3]
base_data_pivot = base_data_pivot[exst_cols + ['Count of TBMs','Count of Unique HCPs Participating','Count of Requests raised']]


base_data_pivot = complete_pivot_columns(base_data_pivot,'TBM Division')

################ Division Level Total ################

summary_rows = []

for i in base_data_pivot['AFFILIATE'].unique():
    base_data_pivot_temp = base_data_pivot[base_data_pivot['AFFILIATE'] == i]

    new_row = {
        'AFFILIATE': f'{i} Total',
        'TBM Division': '999',
        'DIV_NAME': '',
        'Count of TBMs': base_data_pivot_temp['Count of TBMs'].sum(),
        'Count of Unique HCPs Participating': base_data_pivot_temp['Count of Unique HCPs Participating'].sum(),
        'Count of Requests raised': base_data_pivot_temp['Count of Requests raised'].sum(),
        'Count of Requests dispatched': base_data_pivot_temp['Count of Requests dispatched'].sum(),
        'Count of Requests - Action pending / In Process': base_data_pivot_temp['Count of Requests - Action pending / In Process'].sum(),
        'Count of Requests - Dispatched & In Transit': base_data_pivot_temp['Count of Requests - Dispatched & In Transit'].sum(),
        'Count of Requests - Out of stock': base_data_pivot_temp['Count of Requests - Out of stock'].sum(),
        'Count of Requests - Delivered': base_data_pivot_temp['Count of Requests - Delivered'].sum(),
        'Count of Requests - Return': base_data_pivot_temp['Count of Requests - Return'].sum(),
        'Count of Requests - Incomplete Address': base_data_pivot_temp['Count of Requests - Incomplete Address'].sum(),
        'Count of Requests - Dr. Non contactable': base_data_pivot_temp['Count of Requests - Dr. Non contactable'].sum(),
        'Count of Requests - Doctor refused to accept': base_data_pivot_temp['Count of Requests - Doctor refused to accept'].sum()
    }

    summary_rows.append(pd.DataFrame([new_row]))

# Concatenate all summary rows and append to the original DataFrame
summary_df = pd.concat(summary_rows, ignore_index=True)
base_data_pivot = pd.concat([base_data_pivot, summary_df], ignore_index=True)


base_data_pivot = base_data_pivot.sort_values(by=['AFFILIATE','TBM Division'])
base_data_pivot['TBM Division'] = base_data_pivot['TBM Division'].replace('999','')
    

# Initialize Outlook
outlook = win32.Dispatch("Outlook.Application")

#df['ABM Emp Code'] = df['ABM Emp Code'].astype(str)

#df_temp = df.loc[:,['TBM Division','Division Name','Head Management ']].drop_duplicates()

#df_temp = df_temp[(df_temp['Head Management '] != 0)&(df['Head Management '] != '0')]
#df_temp = df_temp[df_temp['Head Management '].notna()]

#base_data_pivot = base_data_pivot[base_data_pivot['TBM Division'].isin([75, 80, 33])]

# Group by Created Alias
grouped = base_data_pivot.groupby("TBM Division") 

for div, group in grouped:
    
    email_df_filt = email_df[email_df['Division Code'] == div]
    
    to_mail = '; '.join(email_df_filt['Email id'].dropna().astype(str).tolist())

    #to_mail = df_temp[df_temp['TBM Division'] == div].iloc[0, 2]
    
    div_name = email_df_filt['Division Name'].unique()[0]
    
    print(div_name)
    
    filename = os.path.join(new_folder_path, f'Sample Request Raised - {div_name}.xlsx')
    #pdfname = os.path.join(input_dir, 'E-Consent Process-Project Reach.pdf')
    #password = 'abbott@123'  # Set your password here
    
    df_group = df[df['TBM Division'] == div].copy()
    
    df_group.rename(columns={'Input Sample Request: Created By':'TBM Name',
                          'Date':'Requested_Date',
                          'Assigned Request Ids':'Request_ID',
                          'Doctor: Customer Code':'Doctor_Code',
                          'Doctor: SAP Customer Code(New)':'SAP_Customer_Code',
                          'Doctor: Account Name':'Doctor_Name',
                          'Item Code':'Item_Code',
                          'SKU':'SKU_Name',
                          'Requested Quantity':'Requested_Quantity',
                          'Request Status':'Request_Status',
                          'Delivery Date':'Delivery Date'}, inplace=True)
    
    df_group = df_group.loc[:,['AFFILIATE','TBM Division','TBM HQ','TBM Name','Requested_Date','Month','Request_ID', 'Doctor_Code', 'SAP_Customer_Code', 'Doctor_Name', 'Item_Code', 'SKU_Name', 'Requested_Quantity', 'Request_Status','Dispatch Date','Delivery Date','Rto Reason']]
    
    #group = group.drop(columns=['email'])
    #group = group.drop(columns=['ABM_email'])
    #group = group.drop(columns=['ZBM_email'])
    
    # Save the DataFrame as an Excel file
    save_excel(df_group, filename)
    
    
    
    #email_id = group["TBM EMAIL_ID"].iloc[0]
    #requested_date = group["Date"].iloc[0]
 
    # Prepare rows data for the template
    rows = []
    for _, row in group.iterrows():
        rows.append({
            "Affiliate": row["AFFILIATE"],
            "TBM_Division": row["TBM Division"],
            "DIV_NAME": row["DIV_NAME"],
            "Count_of_TBMs": row["Count of TBMs"],
            "Count_of_HCPs": row["Count of Unique HCPs Participating"],
            "req_raised": row["Count of Requests raised"],
            "req_dispatched": row["Count of Requests dispatched"],
            "req_pending": row["Count of Requests - Action pending / In Process"],
            "req_in_transit": row["Count of Requests - Dispatched & In Transit"],
            "req_out_of_stock":row["Count of Requests - Out of stock"],
            "req_delivered":row["Count of Requests - Delivered"],
            "req_returned": row["Count of Requests - Return"],
            "req_incomplete_add":row["Count of Requests - Incomplete Address"],
            "req_dr_non_contactable": row["Count of Requests - Dr. Non contactable"],
            "req_dr_refused": row["Count of Requests - Doctor refused to accept"]
        })
 
    # Render HTML email content
    email_html = template.render(rows=rows)
 
    # Create and send email
    mail = outlook.CreateItem(0)
    
    mail.Attachments.Add(filename)
    
    #mail.BCC = 'vaibhav.nalawade@abbott.com'
    aff_name = group['AFFILIATE'].unique().tolist()[0]
    
    mail.To = to_mail
    
    if aff_name == 'AIL':
        mail.cc = 'ishan.mithbavkar@abbott.com;ashwini.suryavanshi@abbott.com;sandesh.bhoir@abbott.com'
    elif aff_name == 'APC':
       mail.cc = 'jenita.nadar@abbott.com;ashwini.suryavanshi@abbott.com;sandesh.bhoir@abbott.com'
    elif aff_name == 'ASC':
        mail.cc = 'sandesh.bhoir@abbott.com;ashwini.suryavanshi@abbott.com'
        
    mail.bCC = 'vaibhav.nalawade@abbott.com;kranti.vengurlekar@abbott.com'
    mail.Subject = f"{div_name}: Sample Direct Dispatch to Doctors - Request Status as of {current_date}"
    mail.HTMLBody = email_html
    mail.SentOnBehalfOfName = 'EPD_SFA@abbott.com'
    mail.Display()  # use mail.Send() to send automatically
  
       
print("All emails sent successfully!")    
    
    
    





